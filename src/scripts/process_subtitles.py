import numpy as np
import pandas as pd
from datasets import load_dataset
import os
from tqdm import tqdm
from transformers import AutoModelForCausalLM, AutoTokenizer, AutoModel
from openpyxl import Workbook, load_workbook
import torch
torch.cuda.empty_cache()

current_dir = os.getcwd()
desired_dir = os.path.dirname(os.path.dirname(current_dir))
prefix_path = desired_dir

def summarize_long_text_with_embeddings(text, tokenizer, model, device):
    """
    Summarize a long text by dividing it into chunks, summarizing each chunk,
    and combining the embedding vectors of the summaries.

    Args:
        text (str): The long text to summarize.
        tokenizer: The tokenizer for text processing.
        model: The summarization model.
        device: The device for model inference (e.g., "cpu" or "cuda").

    Returns:
        str: The concatenated summaries of all chunks.
        torch.Tensor: The combined embedding vector for the entire text.
    """
    summaries = []
    summary_embeddings = []

    tokenized_output = tokenizer(text, return_tensors="pt", add_special_tokens=False)
    all_token_ids = tokenized_output['input_ids'][0]
    
    # Process text in chunks of 1024 tokens
    for i in range(0, len(all_token_ids), 1024):
        chunk_token_ids = all_token_ids[i:i + 1024]
        chunk_text = tokenizer.decode(chunk_token_ids, skip_special_tokens=True)
        
        # Summarize the chunk
        prompt = f"""
        סכם את הטקסט הבא בעברית, זה כתוביות של פרק בסדרה, תתקמד בעלילת הפרק ובדמויות:
        {chunk_text}
        
        תקציר:
        """
        
        # Tokenize the prompt
        with torch.no_grad():
            encoded = tokenizer(prompt, return_tensors="pt").to(device)
        
            # Generate the summary
            generated_ids = model.generate(
                encoded.input_ids,
                max_new_tokens=250,
                do_sample=False,
            )
            
            # Decode the generated summary
            decoded_text = tokenizer.decode(generated_ids[0], skip_special_tokens=True)
            summary = decoded_text.split("תקציר:")[-1].strip()
            summaries.append(summary)
            
            # Get the embedding vector for the summary
            summary_encoded = tokenizer(summary, return_tensors="pt").to(device)
            output = model(**summary_encoded, output_hidden_states=True, return_dict=True)
            summary_embedding = output.hidden_states[-1].mean(dim=1)  # Mean pooling for the summary embedding
            summary_embeddings.append(summary_embedding)

        # Free up GPU memory
        del encoded, generated_ids, summary_encoded, output
        torch.cuda.empty_cache()

    # Concatenate all summaries
    full_summary = " ".join(summaries)
    
    # Combine all summary embeddings using mean pooling
    combined_embedding = torch.mean(torch.stack(summary_embeddings), dim=0)
    
    return full_summary, combined_embedding

def process_and_save_in_batches(df, tokenizer, model, device, output_path, batch_size=1000):
    """
    Process a DataFrame to summarize subtitles and save the results in batches to an Excel file.

    Args:
        df (pd.DataFrame): The input DataFrame with 'sub_id' and 'subtitles' columns.
        tokenizer: The tokenizer for the summarization model.
        model: The summarization model.
        device: The device for model inference (e.g., "cpu" or "cuda").
        output_path (str): The path to save the Excel file.
        batch_size (int): Number of rows to process before saving results.

    Returns:
        None
    """
    # Check if the Excel file exists and load it
    try:
        workbook = load_workbook(output_path)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['sub_id', 'full_summary', 'combined_embedding'])  # Write headers

    processed_data = []

    for index, row in tqdm(df.iterrows(), total=len(df), desc="Processing"):
        sub_id = row['sub_id']
        subtitles = row['subtitles']

        try:
            # Run the subtitles through the summarization function
            full_summary, combined_embedding = summarize_long_text_with_embeddings(
                text=subtitles,
                tokenizer=tokenizer,
                model=model,
                device=device
            )

            # Convert the embedding tensor to a string for saving
            combined_embedding_str = ','.join(map(str, combined_embedding.cpu().numpy().flatten()))

            # Append to processed_data
            processed_data.append([sub_id, full_summary, combined_embedding_str])

        except Exception as e:
            print(f"Error processing record {index + 1} (sub_id: {sub_id}): {e}")
            continue

        # Save batch results to the Excel file
        if len(processed_data) >= batch_size:
            for record in processed_data:
                sheet.append(record)
            workbook.save(output_path)  # Save to disk
            processed_data.clear()  # Clear batch

    # Save any remaining data
    if processed_data:
        for record in processed_data:
            sheet.append(record)
        workbook.save(output_path)

    print(f"Processing complete. Results saved to {output_path}")

if __name__ == "__main__":
    # Load your data
    input_file = "imdb_text_subtitles.csv"
    output_file = prefix_path + "/data/summarized_subtitles.xlsx"

    # Load the DataFrame
    df_imdb_sub = pd.read_csv(os.path.join(prefix_path, "data", input_file))

    # Load already processed sub_ids from the Excel file
    processed_ids = set()
    try:
        workbook = load_workbook(output_file)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):  # Skip header
            processed_ids.add(row[0])  # Assuming sub_id is in the first column
    except FileNotFoundError:
        print("Output file does not exist. Starting fresh.")

    # Filter out already processed rows
    df_to_process = df_imdb_sub[~df_imdb_sub['sub_id'].isin(processed_ids)]

    if df_to_process.empty:
        print("All rows have been processed. Nothing to do!")
    else:
        print(f"Rows remaining to process: {len(df_to_process)}")

        # Initialize model and tokenizer
        device = "cuda" if torch.cuda.is_available() else "cpu"
        model_name = "dicta-il/dictalm2.0-instruct"

        model = AutoModelForCausalLM.from_pretrained(
            model_name,
            torch_dtype=torch.float16,
            device_map="cuda" if device == "cuda" else None
        )
        tokenizer = AutoTokenizer.from_pretrained(model_name)

        # Ensure the model is on the correct device
        model.to(device)

        # Process the data and save results
        process_and_save_in_batches(df_to_process, tokenizer, model, device, output_file, batch_size=100)







